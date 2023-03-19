import openpyxl


file_excel_A = "test_data\\file_example_XLSX_100.xlsx"
workbook_A = openpyxl.load_workbook(file_excel_A)
sheet_A = workbook_A["Sheet1"]
number_of_rows = sheet_A.max_row  # 101
number_of_columns = sheet_A.max_column  # 8

print(f'Sheet max row: {number_of_rows}')
print(f'Sheet max column: {number_of_columns}')

for row in range(1, number_of_rows + 1):
    for column in range(1, number_of_columns + 1):
        print(sheet_A.cell(row, column).value, end='  ')
    print()

file_excel_B = "test_data\\file_test_excel.xlsx"
workbook_B = openpyxl.load_workbook(file_excel_B)
sheet_B = workbook_B.active  # get active sheet from excel

for row in range(1, 10):
    for column in range(1, 5):
        sheet_B.cell(row, column).value = f"I am at {row}:{column}"

workbook_B.save(file_excel_B)

